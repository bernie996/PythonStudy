""""
Module openpyxl

Used to manipulate excel files

--------------------------
Terminology
--------------------------

Spreadsheet/Workbook: Is the file that is being created/used to work with
Sheet/Worksheet: A partition of the Workbook, used to allocate different kinds of content in the same file
Column: Vertical line used to separate. Represented by an uppercase
Rows: Horizontal line used to separate. Represented by a number
Cell: It's a representation of a space delimited by a column and a row. It uses both uppercase from column and number from the row

----------------------------

Instalation

python -m pip install openpyxl

------------------------------

Creation of a Workbook:

name_workbook = openpyxl.Workbook()

-------------------------------

Activation of a Sheet

name_sheet = name_workbook.active

--------------------------------

Edit a Cell by Coordinates

name_sheet["A1"] = "Mamma "
name_sheet["B1"] = "Mia"

--------------------------------

Reading an Excel Spreadsheet

name_workbook = openpyxl.load_workbook(filename="file.xlsx")

--------------------------------

Getting all Sheets availables from a Workplace

name_workbook.sheetnames

--------------------------------

Select a Sheet to work from a Workplace

sheet1 = name_workbook.active

By default, it selects the first available one

-------------------------------

Retrieve Data from a Sheet

sheet1["A1"] # Return a element Cell of the type: # <Cell "Sheet 1" .A1>
sheet["A1"].value # Return the value associated to the cell # "Value"

    It can also be done using the cell() method

sheet.cell(row=1, column=1) # Return a element Cell of the type: # <Cell "Sheet 1" .A1>
sheet.cell(row=1, column=1).value # Return the value associated to the cell # "Value"

    Note: It uses a start-in-one notation, starting both rows and columns in 1

    Options

read_only: Only allows to read the data on the cells, but also allow to open very large files
data_only: Retrieve the result value of a cell, ignoring formulas

------------------------------
Data Iteration

Methods to manage large amounts of data. The return will be a Tuple of elements

    Slicing

Column:
sheet1["A"] #Return all elements from the column A
sheet1["A:C"] #Return all elements of the columns from A to C

Rows:
sheet1["1"] #Return all elements from the row 1
sheet1["1:3"] #Return all elements of the rows from 1 to 3

Cells(both)
sheet1["A1:B2"] # Return the cells from A1 to B2 #
sheet1["A1:B2"].value # Return the value of the cells from A1 to B2 #

------------------------------
Python Generators



    Methods

.iter_rows() The result will be a tuple element per row selected
.iter_cols() The result will be a tuple element per column selected

    Arguments

min_row
max_row
min_col
max_col
values_only (boolean)

Example:

for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3):
    print(row)
(<Cell 'Sheet 1'.A1>, <Cell 'Sheet 1'.B1>, <Cell 'Sheet 1'.C1>)
(<Cell 'Sheet 1'.A2>, <Cell 'Sheet 1'.B2>, <Cell 'Sheet 1'.C2>)

for column in sheet.iter_cols(min_row=1, max_row=2, min_col=1, max_col=3):
    print(column)
(<Cell 'Sheet 1'.A1>, <Cell 'Sheet 1'.A2>)
(<Cell 'Sheet 1'.B1>, <Cell 'Sheet 1'.B2>)
(<Cell 'Sheet 1'.C1>, <Cell 'Sheet 1'.C2>)

    Iterate through the entire dataset

Attributes:

.iter.rows
.iter.columns

Without any arguments both represent shortcuts of the methods .iter.rows() and .iter.cols()


------------------------------

------------------------------

------------------------------

"""

"""
Example - Create
-------------------------------
import openpyxl
from openpyxl import Workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet["A1"] = "Hello "
sheet["A2"] = "World"
workbook.save(filename="hello.xlsx")


"""

"""
from openpyxl import load_workbook
workbook = load_workbook(filename="sample.xlsx")
workbook.sheetnames
"""