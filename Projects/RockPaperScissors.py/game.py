import os
import random
import openpyxl
try:
    historial = openpyxl.load_workbook(filename='historial.xlsx')
except:
    historial = openpyxl.Workbook()
    used = historial.active
    used["A1"] = "Match"
    used["B1"] = "First Participant"
    used["C1"] = "Election"
    used["D1"] = "Second Participant"
    used["E1"] = "Election"
    used["F1"] = "Result"
    historial.save(filename="historial.xlsx")
used = historial.active
def clear(): return os.system('cls')
option1 = 0
sele1 = 0
sele2 = 0
historial_elements = 0
elements = {1: "Rock", 2:"Paper",3:"Scissors"}
print("User, welcome to the Rock, Paper Scissors game. \nPlease select how you want to play:")
while option1 != 4:
    clear()
    option1 = input('(1)Single Player \n(2)Multiplayer \n(3)Matches \n(4)Exit \nOption: ')
    while option1.isdigit() == False or any([ 1 if n == int(option1) else 0 for n in range(1,5) ]) == False:
        clear()
        option1 = input('Please enter a valid option \n(1)Single Player \n(2)Multiplayer \n(3)Vs IA \n(4)Exit \nOption: ')
    option1 = int(option1)
    if option1 == 1:
        option2 = 0
        while option2 != 2:
            counter1=0
            counter2=0
            clear()
            game_mode = input("Which game condition you preffer?\n(1)Single Play\n(2)Best Of \nOption: ")
            while game_mode.isdigit() ==False or  any([1 if n == int(game_mode) else 0 for n in range(1,3)]) == False:
                clear()
                print('Please select a valid option:')
                game_mode = input("Which game condition you preffer?\n(1)Single Play\n(2)Best Of \nOption: ")
            game_mode = int(game_mode)
            if game_mode == 2:
                while True:
                    try:
                        best_of = input("Select the Amount \nBest Of: ")
                        game_mode = int(best_of)
                        break
                    except ValueError:
                        print("Please enter a valid number\n")
            clear()
            for n in range(game_mode):
                sele1 = input('(1)Rock \n(2)Paper \n(3)Scissors \n\nOption: ')
                while sele1.isdigit() == False or any([1 if n == int(sele1) else 0 for n in range(1,4)]) == False:
                    print('Please select a valid option:')
                    sele1 = input('(1)Rock \n(2)Paper \n(3)Scissors \n\nOption: ')
                sele1 = int(sele1)
                sele2 = random.randint(1,3)
                historial_elements += 1
                used.cell(row= historial_elements + 1, column=1).value = historial_elements
                used.cell(row= historial_elements + 1, column=2).value = "User"
                used.cell(row= historial_elements + 1, column=4).value = "PC"
                print("\nYour selection: ",elements[sele1],"\nIA selection: ",elements[sele2])
                if sele1 == 1:
                    used.cell(row= historial_elements + 1, column=3).value = "Rock"
                    if sele2 == 1:
                        print("It's a tie, both have the same selection\n")
                        used.cell(row= historial_elements + 1, column=5).value = "Rock"
                        used.cell(row= historial_elements + 1, column=6).value = "Tie"
                    elif sele2== 2:
                        print("IA wins this round, ",elements[sele2]," beats ",elements[sele1],"\n")
                        counter2 += 1
                        used.cell(row= historial_elements + 1, column=5).value = "Paper"
                        used.cell(row= historial_elements + 1, column=6).value = "PC Wins"
                    elif sele2== 3:
                        print("You win this round, ",elements[sele1]," beats ",elements[sele2],"\n")
                        counter1 += 1
                        used.cell(row= historial_elements + 1, column=5).value = "Scissors"
                        used.cell(row= historial_elements + 1, column=6).value = "User Wins"
                if sele1 == 2:
                    used.cell(row= historial_elements + 1, column=3).value = "Paper"
                    if sele2 == 1:
                        print("You win this round, ",elements[sele1]," beats ",elements[sele2],"\n")
                        counter1 += 1
                        used.cell(row= historial_elements + 1, column=5).value = "Rock"
                        used.cell(row= historial_elements + 1, column=6).value = "User Wins"
                    elif sele2== 2:
                        print("It's a tie, both have the same selection\n")
                        used.cell(row= historial_elements + 1, column=5).value = "Paper"
                        used.cell(row= historial_elements + 1, column=6).value = "Tie"
                    elif sele2== 3:
                        print("IA wins this round, ",elements[sele2]," beats ",elements[sele1],"\n")
                        counter2 += 1
                        used.cell(row= historial_elements + 1, column=5).value = "Scissors"
                        used.cell(row= historial_elements + 1, column=6).value = "PC Wins"
                if sele1 == 3:
                    used.cell(row= historial_elements + 1, column=3).value = "Scissors"
                    if sele2 == 1:
                        print("IA wins this round, ",elements[sele2]," beats ",elements[sele1],"\n")
                        counter2 += 1
                        used.cell(row= historial_elements + 1, column=5).value = "Rock"
                        used.cell(row= historial_elements + 1, column=6).value = "PC Wins"
                    elif sele2== 2:
                        print("You win this round, ",elements[sele1]," beats ",elements[sele2],"\n")
                        counter1 += 1
                        used.cell(row= historial_elements + 1, column=5).value = "Paper"
                        used.cell(row= historial_elements + 1, column=6).value = "User Wins"
                    elif sele2== 3:
                        print("It's a tie, both have the same selection\n")
                        used.cell(row= historial_elements + 1, column=5).value = "Scissors"
                        used.cell(row= historial_elements + 1, column=6).value = "It's a tie"
                print('Score: \nYou: ',counter1,'\nIA: ',counter2,"\n")
            if counter1 > counter2:
                print("Final Result: User Wins\n")
            elif counter1 < counter2:
                print("Final Result: CPU Wins\n")
            else:
                print("Final Result: It's a tie\n")
            option2 = input('What do you want to do now? \n(1)Repeat \n(2)Return to Main Menu \nOption: ')
            while option2.isdigit() == False or any([1 if n == int(option2) else 0 for n in range(1,3)]) == False:
                clear()
                print('\nPlease select a valid option:')
                option2 = input('(1)Repeat \n(2)Return to Main Menu \nOption: ')
            option2 = int(option2)
            historial.save(filename="historial.xlsx")
    if option1 == 2:
        option2 = 0
        while option2 != 2:
            counter1=0
            counter2=0
            clear()
            game_mode = input("Which game condition you preffer?\n(1)Single Play\n(2)Best Of \nOption: ")
            while game_mode.isdigit() ==False or  any([1 if n == int(game_mode) else 0 for n in range(1,3)]) == False:
                clear()
                print('Please select a valid option:')
                game_mode = input("Which game condition you preffer?\n(1)Single Play\n(2)Best Of \nOption: ")
            game_mode = int(game_mode)
            if game_mode == 2:
                    while True:
                        try:
                            best_of = input("Select the Amount \nBest Of: ")
                            game_mode = int(best_of)
                            break
                        except ValueError:
                            print("Please enter a valid number\n")
            clear()
            for n in range(game_mode):
                sele1 = input('(1)Rock \n(2)Paper \n(3)Scissors \n\nOption: ')
                while sele1.isdigit() == False or any([1 if n == int(sele1) else 0 for n in range(1,4)]) == False:
                    print('Please select a valid option:')
                    sele1 = input('(1)Rock \n(2)Paper \n(3)Scissors \n\nOption: ')
                sele1 = int(sele1)
                clear()
                sele2 = input('(1)Rock \n(2)Paper \n(3)Scissors \n\nOption: ')
                while sele2.isdigit() == False or any([1 if n == int(sele2) else 0 for n in range(1,4)]) == False:
                    print('Please select a valid option:')
                    sele2 = input('(1)Rock \n(2)Paper \n(3)Scissors \n\nOption: ')
                sele2 = int(sele2)
                historial_elements += 1
                clear()
                used.cell(row= historial_elements + 1, column=1).value = historial_elements
                used.cell(row= historial_elements + 1, column=2).value = "User 1"
                used.cell(row= historial_elements + 1, column=4).value = "User 2"
                print("\nYour selection: ",elements[sele1],"\nIA selection: ",elements[sele2])
                if sele1 == 1:
                    used.cell(row= historial_elements + 1, column=3).value = "Rock"
                    if sele2 == 1:
                        print("It's a tie, both have the same selection\n")
                        used.cell(row= historial_elements + 1, column=5).value = "Rock"
                        used.cell(row= historial_elements + 1, column=6).value = "Tie"
                    elif sele2== 2:
                        print("User 2 win this round, ",elements[sele2]," beats ",elements[sele1],"\n")
                        counter2 += 1
                        used.cell(row= historial_elements + 1, column=5).value = "Paper"
                        used.cell(row= historial_elements + 1, column=6).value = "PC Wins"
                    elif sele2== 3:
                        print("User 1 win this round, ",elements[sele1]," beats ",elements[sele2],"\n")
                        counter1 += 1
                        used.cell(row= historial_elements + 1, column=5).value = "Scissors"
                        used.cell(row= historial_elements + 1, column=6).value = "User Wins"
                if sele1 == 2:
                    used.cell(row= historial_elements + 1, column=3).value = "Paper"
                    if sele2 == 1:
                        print("User 1 win this round, ",elements[sele1]," beats ",elements[sele2],"\n")
                        counter1 += 1
                        used.cell(row= historial_elements + 1, column=5).value = "Rock"
                        used.cell(row= historial_elements + 1, column=6).value = "User Wins"
                    elif sele2== 2:
                        print("It's a tie, both have the same selection\n")
                        used.cell(row= historial_elements + 1, column=5).value = "Paper"
                        used.cell(row= historial_elements + 1, column=6).value = "Tie"
                    elif sele2== 3:
                        print("User 2 win this round, ",elements[sele2]," beats ",elements[sele1],"\n")
                        counter2 += 1
                        used.cell(row= historial_elements + 1, column=5).value = "Scissors"
                        used.cell(row= historial_elements + 1, column=6).value = "PC Wins"
                if sele1 == 3:
                    used.cell(row= historial_elements + 1, column=3).value = "Scissors"
                    if sele2 == 1:
                        print("User 2 win this round, ",elements[sele2]," beats ",elements[sele1],"\n")
                        counter2 += 1
                        used.cell(row= historial_elements + 1, column=5).value = "Rock"
                        used.cell(row= historial_elements + 1, column=6).value = "PC Wins"
                    elif sele2== 2:
                        print("User 1 win this round, ",elements[sele1]," beats ",elements[sele2],"\n")
                        counter1 += 1
                        used.cell(row= historial_elements + 1, column=5).value = "Paper"
                        used.cell(row= historial_elements + 1, column=6).value = "User Wins"
                    elif sele2== 3:
                        print("It's a tie, both have the same selection\n")
                        used.cell(row= historial_elements + 1, column=5).value = "Scissors"
                        used.cell(row= historial_elements + 1, column=6).value = "It's a tie"
                print('Score: \nUser 1: ',counter1,'\nUser 2: ',counter2,"\n")
            if counter1 > counter2:
                print("Final Result: User 1 Wins\n")
            elif counter1 < counter2:
                print("Final Result: User 2 Wins\n")
            else:
                print("Final Result: It's a tie\n")
            option2 = input('What do you want to do now? \n(1)Repeat \n(2)Return to Main Menu \nOption: ')
            while option2.isdigit() == False or any([1 if n == int(option2) else 0 for n in range(1,3)]) == False:
                clear()
                print('\nPlease select a valid option:')
                option2 = input('(1)Repeat \n(2)Return to Main Menu \nOption: ')
            option2 = int(option2)
            historial.save(filename="historial.xlsx")
    if option1 == 3:
        option2 = 0
        if historial_elements == 1:
            print ("Historial is empyty")
            input("Introduce any key to return to main menu")
        else:
            print(used["A1"].value,"|",used["B1"].value,"|",used["C1"].value,"|",used["D1"].value,"|",used["E1"].value,"|",used["F1"].value)
            for n in used.iter_rows(min_row = 2, values_only = True):
                print("{0} || {1}: {2} vs {3}: {4} || {5}".format(n[0],n[1],n[2],n[3],n[4],n[5]))
            option2 = input('What do you want to do now? \n(1)Delete History \n(2)Return to Main Menu \nOption: ')
            while option2.isdigit() == False or any([1 if n == int(option2) else 0 for n in range(1,3)]) == False:
                clear()
                print('\nPlease select a valid option:')
                option2 = input('(1)Delete History \n(2)Return to Main Menu \nOption: ')
            option2 = int(option2)
            if option2 == 1:
                used.delete_rows(idx= 2, amount = historial_elements) # Hallar la forma de hacer un contador
                print("The historial has been deleted")
                input("Introduce any key to return to main menu")
                historial.save(filename="historial.xlsx")
print("Good bye user!")
